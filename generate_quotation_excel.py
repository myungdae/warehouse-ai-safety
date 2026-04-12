"""
견적서 Excel 생성 스크립트
Warehouse AI Drone System — QT-2026-0409-SW-001
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

import os

wb = openpyxl.Workbook()

# ─────────────────────────────────────────
# 공통 색상 팔레트
# ─────────────────────────────────────────
C_DARK       = "0F172A"   # 커버 배경 (네이비 블랙)
C_NAVY       = "1E293B"
C_INDIGO     = "4F46E5"
C_INDIGO2    = "6366F1"
C_CYAN       = "06B6D4"
C_EMERALD    = "10B981"
C_AMBER      = "F59E0B"
C_ROSE       = "F43F5E"
C_VIOLET     = "8B5CF6"
C_WHITE      = "FFFFFF"
C_LIGHT_BG   = "F8FAFC"
C_BORDER     = "E2E8F0"
C_TEXT_DARK  = "1E293B"
C_TEXT_MID   = "475569"
C_TEXT_LIGHT = "94A3B8"
C_CAT_BG     = "EEF2FF"   # 카테고리 행 배경
C_CAT_FG     = "4338CA"
C_SUB_BG     = "F1F5F9"
C_GRAND_BG   = "0F172A"
C_GRAND_FG   = "34D399"

# ─────────────────────────────────────────
# 헬퍼 함수
# ─────────────────────────────────────────
def fill_solid(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border_thin(color="CBD5E1"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium(color="94A3B8"):
    m = Side(style="medium", color=color)
    t = Side(style="thin", color="CBD5E1")
    return Border(left=m, right=m, top=t, bottom=t)

def border_top_thick(color="4F46E5"):
    thick = Side(style="medium", color=color)
    thin  = Side(style="thin", color="E2E8F0")
    return Border(top=thick, left=thin, right=thin, bottom=thin)

def font(name="맑은 고딕", size=10, bold=False, color="1E293B", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def money(n):
    return f"{n:,}"

def set_col_width(ws, col_widths):
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

def apply_outer_border(ws, min_row, max_row, min_col, max_col, color="4F46E5"):
    thick = Side(style="medium", color=color)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            top    = thick if cell.row == min_row else cell.border.top
            bottom = thick if cell.row == max_row else cell.border.bottom
            left   = thick if cell.column == min_col else cell.border.left
            right  = thick if cell.column == max_col else cell.border.right
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)

# ═══════════════════════════════════════════════════════
#  시트 1 — 표지 (Cover)
# ═══════════════════════════════════════════════════════
ws_cover = wb.active
ws_cover.title = "📋 표지"
ws_cover.sheet_view.showGridLines = False

set_col_width(ws_cover, {
    "A": 3, "B": 22, "C": 22, "D": 22, "E": 22, "F": 3
})
for r in range(1, 45):
    ws_cover.row_dimensions[r].height = 18

# 커버 배경 — A1:F44
for r in range(1, 45):
    for c in range(1, 7):
        ws_cover.cell(r, c).fill = fill_solid(C_DARK)

# ── 상단 여백
ws_cover.row_dimensions[1].height = 12
ws_cover.row_dimensions[2].height = 28

# ── 타이틀 블록
def cover_write(row, col, val, sz=10, bold=False, color=C_WHITE, h="left", v="center", wrap=False, merge_to=None):
    cell = ws_cover.cell(row, col, val)
    cell.font = Font(name="맑은 고딕", size=sz, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    cell.fill = fill_solid(C_DARK)
    if merge_to:
        ws_cover.merge_cells(f"{get_column_letter(col)}{row}:{get_column_letter(merge_to)}{row}")

# 배지 라인
ws_cover.row_dimensions[3].height = 22
cover_write(3, 2, "◆ Operational Ontology   ◆ Agentic AI   ◆ MCP (Model Context Protocol)   ◆ Zero-Touch Automation",
            sz=9, bold=True, color="A5B4FC", merge_to=5)

# 메인 타이틀
ws_cover.row_dimensions[4].height = 10
ws_cover.row_dimensions[5].height = 34
cover_write(5, 2, "자율 드론 창고 재고관리 AI 통합 소프트웨어 시스템",
            sz=20, bold=True, color=C_WHITE, merge_to=5)

ws_cover.row_dimensions[6].height = 22
cover_write(6, 2, "Samsung Semiconductor Warehouse — Software Development Quotation",
            sz=11, color="94A3B8", merge_to=5)

ws_cover.row_dimensions[7].height = 8

# 견적 번호 / 날짜
ws_cover.row_dimensions[8].height = 20
cover_write(8, 2, "견적 번호: QT-2026-0409-SW-001", sz=10, bold=True, color="CBD5E1")
cover_write(8, 4, "발행일: 2026년 4월 9일  |  유효기간: 30일", sz=10, color="94A3B8", h="right", merge_to=5)

ws_cover.row_dimensions[9].height = 8

# 구분선
ws_cover.row_dimensions[10].height = 3
for c in range(2, 6):
    cell = ws_cover.cell(10, c)
    cell.fill = fill_solid(C_INDIGO)

ws_cover.row_dimensions[11].height = 10

# ── 발주처 / 공급사 정보 박스
def info_box_cover(start_row, label, lines, color_label="67E8F9", color_val="E2E8F0"):
    ws_cover.row_dimensions[start_row].height = 18
    cell = ws_cover.cell(start_row, 2, f"▌ {label}")
    cell.font = Font(name="맑은 고딕", size=9, bold=True, color=color_label)
    cell.alignment = Alignment(vertical="center")
    cell.fill = fill_solid("1a2540")
    for i, line in enumerate(lines):
        r = start_row + 1 + i
        ws_cover.row_dimensions[r].height = 17
        cell = ws_cover.cell(r, 2, line)
        cell.font = Font(name="맑은 고딕", size=10, color=color_val)
        cell.alignment = Alignment(vertical="center")
        cell.fill = fill_solid("1a2540")
        ws_cover.merge_cells(f"B{r}:C{r}")

info_box_cover(12, "발주처 (Client)", [
    "삼성반도체 주식회사",
    "DS부문 제조&인프라총괄 / 스마트팩토리팀"
])
for r in range(12, 16):
    for c in range(2, 4):
        ws_cover.cell(r, c).fill = fill_solid("1a2540")

info_box_cover(12, "공급사 (Vendor)", [
    "Warehouse AI Solutions",
    "Agentic AI · MCP · Ontology 전문 개발사"
])
# 오른쪽 칸에
for i, line in enumerate(["공급사 (Vendor)", "Warehouse AI Solutions", "Agentic AI · MCP · Ontology 전문 개발사"]):
    r = 12 + i
    cell = ws_cover.cell(r, 4, line)
    cell.font = Font(name="맑은 고딕", size=10 if i > 0 else 9, 
                     bold=(i==0), color="67E8F9" if i==0 else "E2E8F0")
    cell.alignment = Alignment(vertical="center")
    cell.fill = fill_solid("1a2540")
    ws_cover.merge_cells(f"D{r}:E{r}")

ws_cover.row_dimensions[16].height = 10

# ── 시스템 스펙 정보
specs = [
    ("창고 규모", "15 Aisle × 20 Rack × 15 Level = 4,500 위치"),
    ("드론 구성", "Tethered Drone 2대 (Dock A: Aisle 1~7, Dock B: Aisle 8~15)"),
    ("스캔 레이어", "Layer 1 + Layer 2 동시 스캔 (반도체 부품 창고)"),
    ("운영 방식", "야간 자율 순찰 22:00 → AI 분석 → 담당자 출근 전 06:00 자동 출력"),
    ("핵심 기술", "Operational Ontology (OWL/RDF) + Agentic AI + MCP 세계 최초 통합"),
]
ws_cover.row_dimensions[17].height = 18
cell = ws_cover.cell(17, 2, "▌ 시스템 스펙 및 적용 환경")
cell.font = Font(name="맑은 고딕", size=9, bold=True, color="6EE7B7")
cell.fill = fill_solid("0d2035")
cell.alignment = Alignment(vertical="center")
ws_cover.merge_cells("B17:E17")

for i, (k, v) in enumerate(specs):
    r = 18 + i
    ws_cover.row_dimensions[r].height = 17
    c1 = ws_cover.cell(r, 2, k)
    c1.font = Font(name="맑은 고딕", size=9, bold=True, color="94A3B8")
    c1.fill = fill_solid("0d2035")
    c1.alignment = Alignment(vertical="center")
    c2 = ws_cover.cell(r, 3, v)
    c2.font = Font(name="맑은 고딕", size=10, color="E2E8F0")
    c2.fill = fill_solid("0d2035")
    c2.alignment = Alignment(vertical="center", wrap_text=True)
    ws_cover.merge_cells(f"C{r}:E{r}")

ws_cover.row_dimensions[24].height = 10

# ── 금액 강조 박스
for c in range(2, 6):
    ws_cover.cell(25, c).fill = fill_solid("0d3320")
for c in range(2, 6):
    ws_cover.cell(26, c).fill = fill_solid("0d3320")
for c in range(2, 6):
    ws_cover.cell(27, c).fill = fill_solid("0d3320")
for c in range(2, 6):
    ws_cover.cell(28, c).fill = fill_solid("0d3320")
for c in range(2, 6):
    ws_cover.cell(29, c).fill = fill_solid("0d3320")

ws_cover.row_dimensions[25].height = 8
ws_cover.row_dimensions[26].height = 22

cell = ws_cover.cell(26, 2, "소프트웨어 개발 총액 (VAT 별도)")
cell.font = Font(name="맑은 고딕", size=9, bold=True, color="6EE7B7")
cell.fill = fill_solid("0d3320")
cell.alignment = Alignment(vertical="center")
ws_cover.merge_cells("B26:C26")

cell = ws_cover.cell(26, 4, "하드웨어(드론·프린터·서버) 비용 미포함")
cell.font = Font(name="맑은 고딕", size=9, color="475569", italic=True)
cell.fill = fill_solid("0d3320")
cell.alignment = Alignment(vertical="center", horizontal="right")
ws_cover.merge_cells("D26:E26")

ws_cover.row_dimensions[27].height = 38
cell = ws_cover.cell(27, 2, "₩  80,000,000")
cell.font = Font(name="맑은 고딕", size=28, bold=True, color="34D399")
cell.fill = fill_solid("0d3320")
cell.alignment = Alignment(vertical="center")
ws_cover.merge_cells("B27:C27")

cell = ws_cover.cell(27, 4, "VAT 포함 시: ₩ 88,000,000")
cell.font = Font(name="맑은 고딕", size=11, bold=True, color="94A3B8")
cell.fill = fill_solid("0d3320")
cell.alignment = Alignment(vertical="center", horizontal="right")
ws_cover.merge_cells("D27:E27")

ws_cover.row_dimensions[28].height = 18
cell = ws_cover.cell(28, 2, "납부 방식: 계약금 30% → 중도금 30% → 중도금 20% → 잔금 20%  |  개발 기간: 16주 (4개월)")
cell.font = Font(name="맑은 고딕", size=9, color="64748B")
cell.fill = fill_solid("0d3320")
cell.alignment = Alignment(vertical="center")
ws_cover.merge_cells("B28:E28")

ws_cover.row_dimensions[29].height = 8

# 구분선
ws_cover.row_dimensions[30].height = 3
for c in range(2, 6):
    ws_cover.cell(30, c).fill = fill_solid(C_EMERALD)

# 푸터
ws_cover.row_dimensions[31].height = 10
ws_cover.row_dimensions[32].height = 18
cell = ws_cover.cell(32, 2, "Warehouse AI Solutions  |  contact@warehouse-ai.kr  |  ☎ 02-0000-0000  |  warehouse.exko.kr")
cell.font = Font(name="맑은 고딕", size=9, color="334155")
cell.fill = fill_solid(C_DARK)
cell.alignment = Alignment(vertical="center")
ws_cover.merge_cells("B32:E32")

ws_cover.row_dimensions[33].height = 16
cell = ws_cover.cell(33, 2, "⛔ CONFIDENTIAL — 본 문서는 영업 비밀을 포함합니다. 무단 배포·복제를 금합니다.")
cell.font = Font(name="맑은 고딕", size=9, bold=True, color="475569")
cell.fill = fill_solid(C_DARK)
cell.alignment = Alignment(vertical="center")
ws_cover.merge_cells("B33:E33")


# ═══════════════════════════════════════════════════════
#  시트 2 — 상세 견적 (Main Quotation)
# ═══════════════════════════════════════════════════════
ws = wb.create_sheet("📊 상세 견적")
ws.sheet_view.showGridLines = False

set_col_width(ws, {
    "A": 6,    # 번호
    "B": 40,   # 항목명
    "C": 55,   # 상세 내용
    "D": 8,    # 단위
    "E": 8,    # 수량
    "F": 16,   # 단가
    "G": 16,   # 금액
})

# ── 헤더 배너
ws.row_dimensions[1].height = 10
ws.row_dimensions[2].height = 36
ws.merge_cells("A2:G2")
cell = ws.cell(2, 1, "소프트웨어 개발 상세 견적서  |  QT-2026-0409-SW-001  |  ₩ 80,000,000 (VAT 별도)")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color=C_WHITE)
cell.fill = fill_solid(C_DARK)
cell.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[3].height = 22
ws.merge_cells("A3:G3")
cell = ws.cell(3, 1, "삼성반도체 창고 / Operational Ontology + Agentic AI + MCP — 세계 최초 통합 아키텍처 / 43개 항목 / 11개 카테고리")
cell.font = Font(name="맑은 고딕", size=9, color="94A3B8")
cell.fill = fill_solid("1E293B")
cell.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[4].height = 8

# ── 컬럼 헤더
ws.row_dimensions[5].height = 30
headers = ["번호", "개발 항목명", "상세 내용 및 기술 스택", "단위", "수량", "단가 (원)", "금액 (원)"]
header_colors = [C_DARK]*7
for col, (h, bg) in enumerate(zip(headers, header_colors), 1):
    cell = ws.cell(5, col, h)
    cell.font = Font(name="맑은 고딕", size=10, bold=True, color=C_WHITE)
    cell.fill = fill_solid(C_DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_thin("334155")

# ── 데이터 정의
# (번호, 항목명, 상세내용, 단위, 수량, 단가, 금액, is_category, is_subtotal)
rows = [
  # ── A
  ("A", "A. Operational Ontology 설계 · 구축", "OWL/RDF 기반 창고 도메인 지식 그래프 모델링", "", "", "", "", True, False),
  ("A1", "창고 도메인 온톨로지 설계 (OWL/RDF)", "15A×20R×15L 공간 지식 그래프 설계 | 드론·Dock·SKU·Shelf 개념 체계 형식화 | OWL 2.0 / RDF/Turtle / SPARQL 1.1 / Protégé", "식", 1, 4500000, 4500000, False, False),
  ("A2", "ERP 시맨틱 매핑 엔진 개발", "삼성 ERP 스키마 → 온톨로지 자동 매핑 변환기 | Day1/Day2 정합성 추론 | Python / RDFLib / SPARQLWrapper / JSON-LD", "식", 1, 3500000, 3500000, False, False),
  ("A3", "물리 공간 디지털 트윈 데이터 모델", "창고 3D 위상 데이터 구조 설계 | 선반 ID 체계 확립 | 좌표계 수학 모델 | GeoJSON / JSON Schema", "식", 1, 2000000, 2000000, False, False),
  ("A4", "SKU 분류 체계 및 이상 패턴 온톨로지", "반도체 부품 SKU 계층 분류 온톨로지 | 분실·감소·증가 이상 패턴 규칙 | OWL Restrictions / SWRL Rules", "식", 1, 2000000, 2000000, False, False),
  ("A5", "온톨로지 서빙 API 엔드포인트", "/ontology/*.json 정적 서빙 | CORS 설정 | 프론트엔드 실시간 온톨로지 참조 지원", "식", 1, 800000, 800000, False, False),
  ("", "A. 소계", "", "", "", "", 12800000, False, True),

  # ── B
  ("B", "B. Agentic AI 엔진 개발", "자율 의사결정·이상 감지·판정·권고 AI 모듈 개발", "", "", "", "", True, False),
  ("B1", "자율 드론 순찰 의사결정 에이전트", "Dock 출발→Aisle 진입→Rack 스캔→귀환 자율 경로 생성 | 직선 경로 강제 제약(케이블 보호) | Path Algorithm / Canvas API / rAF", "식", 1, 5000000, 5000000, False, False),
  ("B2", "재고 이상 감지 및 자동 판정 AI 모듈", "Day1 vs Day2 재고 차이 분석 | 분실·감소·증가·이동 4분류 | HIGH/MEDIUM/LOW 심각도 판정 | Diff Engine / 통계분석 / Rule-based AI", "식", 1, 4500000, 4500000, False, False),
  ("B3", "자율 Rescan 우선순위 결정 에이전트", "이상 위치 자동 감지 후 Rescan 작업 큐 자율 구성 | Dock별 드론 배정 최적화 | Task Queue / Priority Algorithm", "식", 1, 3500000, 3500000, False, False),
  ("B4", "Rescan Before/After 비교 판정 엔진", "재스캔 전·후 스냅샷 자동 캡처 비교 | ✅ all_found / ⚠️ partial_found / 🚨 confirmed_missing 3단계 판정 | Snapshot Diff / Verdict Engine", "식", 1, 3000000, 3000000, False, False),
  ("B5", "AI 권고사항 자동 생성 엔진 (NLG)", "정확도·분실수·재스캔결과 기반 자연어 권고 자동 생성 | 🔴긴급/🟡주의/🟠권고/🟢양호 4단계 색상 분류 | NLG / Rule Engine / Context Aware", "식", 1, 2500000, 2500000, False, False),
  ("B6", "주요 이상 SKU 자동 랭킹 분석 모듈", "손실 수량 기준 Top-5 이상 SKU 자동 정렬 | 분실·감소·총손실수량 종합 리스크 점수 산출 | Analytics / Risk Scoring / Ranking", "식", 1, 1500000, 1500000, False, False),
  ("B7", "드론 운영 현황 자동 집계 모듈", "Drone A/B 별 스캔 건수·담당 Zone·전력 상태·완료율 자동 집계 | 보고서 드론 현황 섹션 데이터 생성 | Stats Aggregation / Telemetry", "식", 1, 1200000, 1200000, False, False),
  ("", "B. 소계", "", "", "", "", 21200000, False, True),

  # ── C
  ("C", "C. MCP 야간 자동화 시스템", "Model Context Protocol 기반 Zero-Touch 야간 오케스트레이터", "", "", "", "", True, False),
  ("C1", "MCP 야간 스케줄러 백그라운드 데몬", "30초 tick 기반 Python 백그라운드 스케줄러 | 22:00 야간 순찰 자동 트리거 | 06:00 자동 출력 트리거 | Python / Threading / Daemon / Cron Logic", "식", 1, 3000000, 3000000, False, False),
  ("C2", "네트워크 프린터 자동 출력 엔진", "RAW TCP(9100) + IPP(631) 이중 프로토콜 자동 시도 | PJL/ASCII 보고서 포맷 생성 | Python socket 직접 전송 | 오프라인 시 로그·재시도 큐 | RAW TCP / IPP / PJL / Socket", "식", 1, 3500000, 3500000, False, False),
  ("C3", "프린터 온라인 상태 체크 모듈", "출력 전 TCP 연결로 프린터 온라인 확인(타임아웃 3초) | RAW 실패 시 IPP 폴백 | 오프라인 이벤트 로그 | TCP Probe / Health Check", "식", 1, 1000000, 1000000, False, False),
  ("C4", "MCP 설정 관리 시스템 (REST API)", "출력 시각·프린터 IP·순찰 시각·복사 매수·ON/OFF 설정 | JSON 파일 영속 저장 | GET/POST /api/mcp/config | REST API / JSON Config / Dynamic Config", "식", 1, 1500000, 1500000, False, False),
  ("C5", "MCP 이벤트 로그 시스템", "출력 성공·실패·오프라인·순찰 완료 이벤트 로그(최근 200건) | 타임스탬프·트리거 유형·상세 메시지 기록 | GET /api/mcp/log | Event Log / JSON Persistence", "식", 1, 1000000, 1000000, False, False),
  ("C6", "MCP Auto-Print 관리 대시보드 UI", "프린터 상태 실시간 카드 | 스케줄 타임라인 시각화 | 출력 미리보기 팝업 | 수동 출력 버튼 | 설정 폼 | HTML5 / CSS3 / Vanilla JS", "식", 1, 2000000, 2000000, False, False),
  ("C7", "수동 출력 API 및 미리보기 엔드포인트", "POST /api/print/manual — 즉시 수동 출력 트리거 | POST /api/print/preview — 출력 내용 사전 확인 | Flask / REST API / Print Preview", "식", 1, 1000000, 1000000, False, False),
  ("C8", "야간 순찰 신호 브릿지 API", "POST /api/mcp/patrol-signal — MCP → 프론트 드론 UI 트리거 | patrol_trigger.json 파일 기반 신호 전달 | Signal Bridge / JSON Flag / Polling", "식", 1, 800000, 800000, False, False),
  ("", "C. 소계", "", "", "", "", 13800000, False, True),

  # ── D
  ("D", "D. Tethered 드론 제어 & 실시간 창고 시각화 UI", "드론 애니메이션·경로 제어·실시간 SVG 렌더링", "", "", "", "", True, False),
  ("D1", "SVG 실시간 창고 맵 시각화 엔진", "1400×620px SVG 창고 전체 맵 동적 렌더링 | 15 Aisle×2 Side 색상 코딩 | 미스캔(회색)·스캔중(노랑)·완료(초록)·이상(빨강) | SVG / Dynamic Render", "식", 1, 3000000, 3000000, False, False),
  ("D2", "Tethered 드론 2대 실시간 애니메이션", "Dock A/B 출발·Aisle 진입·Rack 스캔·귀환 완전 애니메이션 | Rack-by-Rack 800ms 일시정지 | 속도 3.0 units/frame | JS Animation / rAF / SVG", "식", 1, 3500000, 3500000, False, False),
  ("D3", "Dock 패널 실시간 상태 표시 시스템", "Dock A/B별 드론 ID·담당 Aisle·현재 Aisle·현재 Level 실시간 표시 | L1(녹)·L2(청록) 레벨 색상 코딩 | Real-time Panel / Status Card", "식", 1, 1500000, 1500000, False, False),
  ("D4", "라이브 스캔 피드 로그 시스템", "Rack ID 읽기→Side A/B 스캔→완료 색상 피드(최근 50건 롤링) | System·Rack·Scan·Success·Warning 5종 색상 분류 | Live Feed / Rolling Log", "식", 1, 1000000, 1000000, False, False),
  ("D5", "스캔 커버리지 및 통계 실시간 집계", "총 위치·스캔 건수·커버리지(%)·정확도 실시간 통계 카드 | Day1/Day2 전환 기능 | Statistics / Dynamic UI / Day Switch", "식", 1, 1000000, 1000000, False, False),
  ("D6", "드론 제어 인터럽트 및 Rescan 경로 삽입", "기존 순찰 중단 후 Rescan 작업 경로 맨 앞 삽입 | 재스캔 완료 후 이전 순찰 지점부터 자동 재개 | Path Interrupt / Queue Prepend / Auto Resume", "식", 1, 1500000, 1500000, False, False),
  ("", "D. 소계", "", "", "", "", 11500000, False, True),

  # ── E
  ("E", "E. ERP 연동 · 일일 종합보고서 시스템", "ERP 대조·보고서 자동 생성·Admin 관리 대시보드·다중 내보내기", "", "", "", "", True, False),
  ("E1", "Day1 vs Day2 ERP 비교 분석 뷰", "6-KPI 카드(총위치·분실·감소·증가·정확도·재스캔완료) | 분실·감소·증가 상세 목록 15건+Action Required 섹션 | ERP Integration / Diff View", "식", 1, 2500000, 2500000, False, False),
  ("E2", "일일 종합보고서 자동 생성 엔진", "6섹션(KPI·AI권고·재스캔결과·변동내역·Top5SKU·드론현황) | 보고서 ID 자동 생성(RPT-YYYYMMDD-HHMM) | Report Engine / Auto Generation", "식", 1, 2500000, 2500000, False, False),
  ("E3", "일일 보고서 모달 UI (6개 섹션)", "전체화면 모달 | 헤더 KPI 카드 6종 | AI 권고 색상 코딩 | 재스캔 결과 테이블 | Top-5 SKU 랭킹 | 드론 현황 | Modal UI / 6 Sections", "식", 1, 2000000, 2000000, False, False),
  ("E4", "보고서 Admin 관리 대시보드", "보고서 목록 조회·검색·필터·정렬 대시보드 | 상세 모달·메모 편집(PATCH)·삭제(DELETE)·통계 차트 | Admin Panel / CRUD / Chart.js", "식", 1, 3000000, 3000000, False, False),
  ("E5", "다중 내보내기 시스템", "재스캔 목록 텍스트 내보내기 | 전체 ERP 비교 보고서 다운로드 | JSON 내보내기 | 인쇄 팝업 | ERP API Push 버튼 | Export / JSON / Print API", "식", 1, 1500000, 1500000, False, False),
  ("E6", "이메일 알림 시스템 (SMTP)", "순찰 완료·이상 감지 시 담당자 이메일 자동 발송 | HTML 이메일 템플릿(KPI·분실목록·AI권고 포함) | SMTP / HTML Email / Text Fallback", "식", 1, 1500000, 1500000, False, False),
  ("", "E. 소계", "", "", "", "", 13000000, False, True),

  # ── F
  ("F", "F. 백엔드 서버 & REST API 인프라", "Flask 서버 아키텍처·20개 엔드포인트·SQLite·보안", "", "", "", "", True, False),
  ("F1", "Flask 웹 서버 아키텍처 설계 및 구축", "Flask RESTful 서버 전체 아키텍처 | Blueprint 모듈화 | 정적 파일 최적화 서빙 | 멀티 템플릿 렌더링 | Python / Flask / Blueprint", "식", 1, 2000000, 2000000, False, False),
  ("F2", "전체 REST API 설계 및 개발 — 20개 엔드포인트", "보고서 CRUD · MCP 설정 · 프린터 상태 · 순찰 신호 · 아카이브 등 | GET/POST/PATCH/DELETE 완전 구현 | REST API / 20 Endpoints", "개", 20, 200000, 4000000, False, False),
  ("F3", "SQLite 보고서 아카이브 DB 설계", "영속 SQLite DB 스키마 설계 · 인덱스 최적화 | 검색·필터·정렬 쿼리 최적화 | JSON→SQLite 마이그레이션 스크립트 | SQLite / SQL / Migration", "식", 1, 1500000, 1500000, False, False),
  ("F4", "시스템 보안 및 인증 기본 설계", "API 키 기반 인증 | CORS 정책 설정 | .env 환경변수 분리 | 민감정보 보안 처리 | Security / Auth / CORS / ENV", "식", 1, 1200000, 1200000, False, False),
  ("", "F. 소계", "", "", "", "", 8700000, False, True),

  # ── G
  ("G", "G. Archive Intelligence & 데이터 분석 모듈", "장기 아카이브 분석·통계 API·CSV 내보내기", "", "", "", "", True, False),
  ("G1", "Archive Intelligence 대시보드", "장기 보고서 아카이브 분석 | 날짜별 정확도 트렌드 차트 | 드론별 성능 분석 | 이상 발생 패턴 히트맵 | Archive Analysis / Chart.js", "식", 1, 2000000, 2000000, False, False),
  ("G2", "보고서 통계 API 및 집계 엔진", "GET /api/reports/stats — 정확도 평균·분실 총계·드론별 통계 | 기간별·드론별·정확도 구간별 필터링 | Stats API / Aggregation", "식", 1, 1200000, 1200000, False, False),
  ("G3", "CSV/Excel 보고서 일괄 내보내기", "GET /api/reports/export — 전체 보고서 CSV 일괄 다운로드 | 날짜 범위·드론 필터 | 한글 컬럼명 지원 | Export API / CSV", "식", 1, 800000, 800000, False, False),
  ("", "G. 소계", "", "", "", "", 4000000, False, True),

  # ── H
  ("H", "H. 품질보증 (QA) · 테스트 · 검증", "단위·통합·성능·UAT 테스트 설계 및 실행", "", "", "", "", True, False),
  ("H1", "단위 테스트 설계 및 실행", "핵심 모듈별 단위 테스트 케이스 작성 및 실행 | pytest | AI 판정·스케줄러·API·온톨로지 매핑 검증 | Coverage 80%+", "식", 1, 1500000, 1500000, False, False),
  ("H2", "통합 테스트 및 E2E 시나리오 검증", "야간 순찰→AI 분석→보고서 저장→자동 출력 E2E 검증 | Rescan·ERP 비교·MCP 스케줄러 통합 시나리오 | E2E Test / Integration", "식", 1, 2000000, 2000000, False, False),
  ("H3", "성능 테스트 및 부하 검증", "2대 드론 동시 실행 렌더링 성능 측정 | API 응답 시간 검증 | 24시간 연속 안정성 테스트 | 메모리 누수 점검 | Performance / Load Test", "식", 1, 1000000, 1000000, False, False),
  ("H4", "사용자 수락 테스트 (UAT) 지원", "고객사 담당자 UAT 시나리오 작성 및 현장 지원 | 피드백 수집 및 수정사항 반영 최대 2회 | UAT / On-site / 2 Rounds", "회", 2, 800000, 1600000, False, False),
  ("", "H. 소계", "", "", "", "", 6100000, False, True),

  # ── I
  ("I", "I. 배포 · 운영 지원 · 기술 문서화", "Ubuntu 배포·CI/CD·3개월 운영지원·현장 교육", "", "", "", "", True, False),
  ("I1", "Ubuntu 서버 배포 스크립트 및 자동화", "nohup / PM2 / systemd 배포 구성 스크립트 | Python 의존성·SQLite DB 초기화·환경변수 자동화 | Ubuntu / Bash / systemd / PM2", "식", 1, 1500000, 1500000, False, False),
  ("I2", "GitHub 버전 관리 및 CI/CD 파이프라인", "Git 브랜치 전략 수립 | PR 워크플로우 | 자동 배포 브랜치 스크립트 | GitHub Actions 기반 CI | GitHub / CI/CD / Git Flow", "식", 1, 1000000, 1000000, False, False),
  ("I3", "기술 문서화 (개발자·운영자 매뉴얼)", "API 명세서(Swagger/Markdown) | 시스템 아키텍처 문서 | 운영자 사용 매뉴얼 | MCP 설정 가이드 | API Docs / Architecture / User Manual", "식", 1, 1500000, 1500000, False, False),
  ("I4", "운영 안정화 기술 지원 (3개월)", "납품 후 3개월 버그 수정·운영 질의 대응 | 원격 지원 월 4회 | SLA: 장애 24시간 내 대응 | Support 3M / Remote / SLA", "개월", 3, 1300000, 3900000, False, False),
  ("I5", "현장 설치 및 담당자 교육", "최종 납품 데모 시연 | 담당자 교육 2회(2시간/회) | 현장 서버 설치 지원 | 초기 운영 설정 확인 | On-site / Training", "회", 2, 700000, 1400000, False, False),
  ("", "I. 소계", "", "", "", "", 9300000, False, True),

  # ── J
  ("J", "J. 프로젝트 관리 (PM) & 기획", "요구사항 분석·PM 16주·UI/UX 기획", "", "", "", "", True, False),
  ("J1", "요구사항 분석 및 시스템 아키텍처 설계", "창고 환경 현장 조사·인터뷰·BRD/SRS 문서화 | Ontology·AI·MCP 통합 소프트웨어 아키텍처 완전 설계 | BRD/SRS / Architecture Design", "식", 1, 2000000, 2000000, False, False),
  ("J2", "프로젝트 관리 (PM) — 16주 전기간", "주간 진행 보고·일정 관리·리스크 대응·이슈 트래킹 | 고객사 커뮤니케이션 창구 | 마일스톤 관리 | PM / Agile Scrum", "주", 16, 250000, 4000000, False, False),
  ("J3", "UI/UX 기획 및 프로토타입 설계", "창고 시각화·보고서·MCP 대시보드 UX 흐름 설계 | 와이어프레임 | 인터랙션 정의 | 피드백 반영 | UX Design / Wireframe", "식", 1, 1000000, 1000000, False, False),
  ("", "J. 소계", "", "", "", "", 7000000, False, True),

  # ── K
  ("K", "K. 기술 혁신 프리미엄", "세계 최초 Ontology + Agentic AI + MCP 통합 가치 반영", "", "", "", "", True, False),
  ("K1", "세계 최초 3기술 통합 아키텍처 설계 프리미엄", "Operational Ontology + Agentic AI + MCP 동시 통합 — 업계 전례 없는 구현 | Gartner Hype Cycle 2025/2026 선도 기술 3종 | 재현 불가 독자 아키텍처 IP 가치 | First-Mover / IP Value", "식", 1, 3500000, 3500000, False, False),
  ("K2", "Zero-Touch 야간 자동화 워크플로우 특허 준비", "22:00 순찰→AI 분석→06:00 자동 출력 완전 자동화 플로우 | 기술 차별성 문서화 | 특허 출원 준비 자료 작성 지원 | Patent Ready / Innovation", "식", 1, 3000000, 3000000, False, False),
  ("", "K. 소계", "", "", "", "", 6500000, False, True),
]

# ── 카테고리별 배경색
CAT_FILLS = {
    "A": ("EEF2FF", "4338CA"),  # indigo
    "B": ("E0F2FE", "0369A1"),  # sky
    "C": ("D1FAE5", "065F46"),  # emerald
    "D": ("FEF3C7", "92400E"),  # amber
    "E": ("FCE7F3", "9D174D"),  # pink
    "F": ("F5F3FF", "5B21B6"),  # violet
    "G": ("FFF7ED", "9A3412"),  # orange
    "H": ("F0FDF4", "14532D"),  # green
    "I": ("F0F9FF", "0C4A6E"),  # sky 2
    "J": ("FDF4FF", "6B21A8"),  # purple
    "K": ("FFF1F2", "881337"),  # rose
}

data_row = 6
for row_data in rows:
    num, name, detail, unit, qty, price, amount, is_cat, is_sub = row_data

    ws.row_dimensions[data_row].height = 44 if (not is_cat and not is_sub) else 22

    cat_key = num[0] if num and len(num) >= 1 else ""
    cat_fill, cat_text = CAT_FILLS.get(cat_key, ("F8FAFC", "1E293B"))

    if is_cat:
        # 카테고리 헤더 행
        ws.merge_cells(f"A{data_row}:G{data_row}")
        cell = ws.cell(data_row, 1, f"  {name}")
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color=cat_text)
        cell.fill = fill_solid(cat_fill)
        cell.alignment = Alignment(vertical="center", horizontal="left")
        b = border_top_thick(cat_text)
        cell.border = b
        for c in range(1, 8):
            ws.cell(data_row, c).fill = fill_solid(cat_fill)
        # 왼쪽 두꺼운 강조선
        ws.cell(data_row, 1).border = Border(
            left=Side(style="thick", color=cat_text),
            top=Side(style="medium", color=cat_text),
            bottom=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1")
        )

    elif is_sub:
        # 소계 행
        ws.merge_cells(f"A{data_row}:E{data_row}")
        c_label = ws.cell(data_row, 1, f"  {name}")
        c_label.font = Font(name="맑은 고딕", size=10, bold=True, color="334155")
        c_label.fill = fill_solid("F1F5F9")
        c_label.alignment = Alignment(horizontal="right", vertical="center")
        c_label.border = border_thin()

        # 소계 금액 셀
        ws.merge_cells(f"F{data_row}:G{data_row}")
        c_amt = ws.cell(data_row, 6, amount)
        c_amt.font = Font(name="맑은 고딕", size=11, bold=True, color=C_INDIGO)
        c_amt.fill = fill_solid("F1F5F9")
        c_amt.number_format = '#,##0'
        c_amt.alignment = Alignment(horizontal="right", vertical="center")
        c_amt.border = border_thin()
        for c in range(1, 8):
            ws.cell(data_row, c).fill = fill_solid("F1F5F9")

    else:
        # 일반 항목 행
        row_bg = "FFFFFF" if data_row % 2 == 0 else "FAFBFF"
        vals = [num, name, detail, unit, qty if qty != "" else "", price if price != "" else "", amount if amount != "" else ""]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(data_row, ci, v)
            cell.fill = fill_solid(row_bg)
            cell.border = border_thin()
            if ci == 1:  # 번호
                cell.font = Font(name="맑은 고딕", size=9, bold=True, color=C_TEXT_LIGHT)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 2:  # 항목명
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color=C_TEXT_DARK)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            elif ci == 3:  # 상세
                cell.font = Font(name="맑은 고딕", size=9, color=C_TEXT_MID)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
            elif ci == 4:  # 단위
                cell.font = Font(name="맑은 고딕", size=10, color=C_TEXT_MID)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 5:  # 수량
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color=C_TEXT_DARK)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci == 6:  # 단가
                cell.font = Font(name="맑은 고딕", size=10, color=C_TEXT_MID)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if v != "":
                    cell.number_format = '#,##0'
            elif ci == 7:  # 금액
                cell.font = Font(name="맑은 고딕", size=10, bold=True, color=C_INDIGO)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if v != "":
                    cell.number_format = '#,##0'

    data_row += 1

# ── 합계 구분선
ws.row_dimensions[data_row].height = 6
for c in range(1, 8):
    ws.cell(data_row, c).fill = fill_solid(C_INDIGO)
data_row += 1

# ── 합계 박스
totals = [
    ("공급가액 합계 (소계 합산 원래 금액)", 113900000, False),
    ("볼륨 할인 / 프로모션 조정액", -33900000, True),
    ("★ 최종 공급가액 합계 (VAT 별도)", 80000000, False),
    ("부가가치세 (VAT 10%)", 8000000, False),
    ("◎ 최 종 합 계 (VAT 포 함)", 88000000, False),
]

total_fills = ["1E293B", "1E293B", "0F172A", "1E293B", "0A1628"]
total_fonts = ["94A3B8", "F87171", "A5B4FC", "94A3B8", "34D399"]
total_sizes = [10, 10, 12, 10, 16]
total_bolds = [False, False, True, False, True]

for (label, amt, is_disc), bg, fg, sz, bo in zip(totals, total_fills, total_fonts, total_sizes, total_bolds):
    ws.row_dimensions[data_row].height = 28 if sz >= 14 else 22
    ws.merge_cells(f"A{data_row}:E{data_row}")
    c1 = ws.cell(data_row, 1, f"  {label}")
    c1.font = Font(name="맑은 고딕", size=sz, bold=bo, color=fg)
    c1.fill = fill_solid(bg)
    c1.alignment = Alignment(horizontal="right", vertical="center")

    ws.merge_cells(f"F{data_row}:G{data_row}")
    c2 = ws.cell(data_row, 6, amt)
    c2.font = Font(name="맑은 고딕", size=sz, bold=bo, color=fg)
    c2.fill = fill_solid(bg)
    c2.alignment = Alignment(horizontal="right", vertical="center")
    c2.number_format = '#,##0'
    data_row += 1

# 빈 줄
ws.row_dimensions[data_row].height = 8
data_row += 1

# ── 주석
note_lines = [
    "※ 본 견적은 소프트웨어 개발 비용만 포함합니다. 하드웨어(드론·프린터·서버·네트워크) 비용은 포함되지 않습니다.",
    "※ VAT(부가가치세 10%)는 별도이며, 각 마일스톤 완료 시점에 세금계산서를 발행합니다.",
    "※ 본 견적의 유효기간은 발행일로부터 30일이며, 이후에는 재견적이 필요합니다.",
    "※ 소스코드 저작권은 최종 잔금 납부 완료 후 발주처로 이전됩니다.",
]
for note in note_lines:
    ws.row_dimensions[data_row].height = 16
    ws.merge_cells(f"A{data_row}:G{data_row}")
    cell = ws.cell(data_row, 1, note)
    cell.font = Font(name="맑은 고딕", size=8.5, italic=True, color="64748B")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    data_row += 1


# ═══════════════════════════════════════════════════════
#  시트 3 — 카테고리 요약 (Summary)
# ═══════════════════════════════════════════════════════
ws_sum = wb.create_sheet("📈 카테고리 요약")
ws_sum.sheet_view.showGridLines = False
set_col_width(ws_sum, {"A": 5, "B": 10, "C": 38, "D": 14, "E": 14, "F": 8, "G": 5})
for r in range(1, 60):
    ws_sum.row_dimensions[r].height = 18

# 헤더
ws_sum.row_dimensions[1].height = 10
ws_sum.row_dimensions[2].height = 34
ws_sum.merge_cells("A2:G2")
cell = ws_sum.cell(2, 1, "카테고리별 비용 요약 및 구성비  |  QT-2026-0409-SW-001")
cell.font = Font(name="맑은 고딕", size=14, bold=True, color=C_WHITE)
cell.fill = fill_solid(C_DARK)
cell.alignment = Alignment(horizontal="center", vertical="center")

ws_sum.row_dimensions[3].height = 10

# 컬럼 헤더
ws_sum.row_dimensions[4].height = 26
hdrs = ["", "카테고리", "내용", "개발 항목 수", "금액 (원)", "구성비 (%)", ""]
for ci, h in enumerate(hdrs, 1):
    cell = ws_sum.cell(4, ci, h)
    cell.font = Font(name="맑은 고딕", size=10, bold=True, color=C_WHITE)
    cell.fill = fill_solid(C_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_thin("334155")

# 카테고리 데이터
cat_data = [
    ("A", "Operational Ontology 설계·구축",       5,  12800000, "EEF2FF", "4338CA"),
    ("B", "Agentic AI 엔진 개발",                  7,  21200000, "E0F2FE", "0369A1"),
    ("C", "MCP 야간 자동화 시스템",                8,  13800000, "D1FAE5", "065F46"),
    ("D", "Tethered 드론 제어 & 시각화 UI",        6,  11500000, "FEF3C7", "92400E"),
    ("E", "ERP 연동 · 일일 종합보고서",            6,  13000000, "FCE7F3", "9D174D"),
    ("F", "백엔드 서버 & REST API",                4,   8700000, "F5F3FF", "5B21B6"),
    ("G", "Archive Intelligence & 분석",           3,   4000000, "FFF7ED", "9A3412"),
    ("H", "QA & 테스트",                           4,   6100000, "F0FDF4", "14532D"),
    ("I", "배포 · 운영지원 · 문서화",              5,   9300000, "F0F9FF", "0C4A6E"),
    ("J", "프로젝트 관리 (PM) & 기획",             3,   7000000, "FDF4FF", "6B21A8"),
    ("K", "기술 혁신 프리미엄",                    2,   6500000, "FFF1F2", "881337"),
]
total_amt = sum(r[3] for r in cat_data)
total_items = sum(r[2] for r in cat_data)

r = 5
for cat, desc, items, amt, bg, fg in cat_data:
    pct = round(amt / 80000000 * 100, 1)
    ws_sum.row_dimensions[r].height = 24
    for c in range(1, 8):
        ws_sum.cell(r, c).fill = fill_solid(bg if c not in (1, 7) else "FFFFFF")
        ws_sum.cell(r, c).border = border_thin()
    ws_sum.cell(r, 2, cat).font = Font(name="맑은 고딕", size=11, bold=True, color=fg)
    ws_sum.cell(r, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.cell(r, 3, desc).font = Font(name="맑은 고딕", size=10, bold=True, color=fg)
    ws_sum.cell(r, 3).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_sum.cell(r, 4, items).font = Font(name="맑은 고딕", size=10, color="334155")
    ws_sum.cell(r, 4).alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.cell(r, 5, amt).font = Font(name="맑은 고딕", size=10, bold=True, color=fg)
    ws_sum.cell(r, 5).number_format = '#,##0'
    ws_sum.cell(r, 5).alignment = Alignment(horizontal="right", vertical="center")
    ws_sum.cell(r, 6, pct).font = Font(name="맑은 고딕", size=10, color="475569")
    ws_sum.cell(r, 6).number_format = '0.0"%"'
    ws_sum.cell(r, 6).alignment = Alignment(horizontal="center", vertical="center")
    r += 1

# 합계 행
ws_sum.row_dimensions[r].height = 10
r += 1
ws_sum.row_dimensions[r].height = 28
for c in range(1, 8):
    ws_sum.cell(r, c).fill = fill_solid(C_DARK)
    ws_sum.cell(r, c).border = border_thin("334155")
ws_sum.cell(r, 2, "합계").font = Font(name="맑은 고딕", size=11, bold=True, color=C_WHITE)
ws_sum.cell(r, 2).alignment = Alignment(horizontal="center", vertical="center")
ws_sum.cell(r, 3, "소프트웨어 개발 43개 항목 합계").font = Font(name="맑은 고딕", size=11, bold=True, color="94A3B8")
ws_sum.cell(r, 3).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_sum.cell(r, 4, total_items).font = Font(name="맑은 고딕", size=11, bold=True, color="94A3B8")
ws_sum.cell(r, 4).alignment = Alignment(horizontal="center", vertical="center")

ws_sum.merge_cells(f"E{r}:F{r}")
ws_sum.cell(r, 5, 80000000).font = Font(name="맑은 고딕", size=14, bold=True, color="34D399")
ws_sum.cell(r, 5).number_format = '#,##0'
ws_sum.cell(r, 5).alignment = Alignment(horizontal="right", vertical="center")
r += 2

# VAT 및 최종
vat_rows = [
    ("공급가액 합계 (VAT 별도)", 80000000, "94A3B8"),
    ("부가가치세 10%", 8000000, "94A3B8"),
    ("최 종 합 계 (VAT 포함)", 88000000, "34D399"),
]
for label, amt, color in vat_rows:
    ws_sum.row_dimensions[r].height = 22
    ws_sum.merge_cells(f"B{r}:E{r}")
    ws_sum.cell(r, 2, label).font = Font(name="맑은 고딕", size=11, bold=True, color="475569")
    ws_sum.cell(r, 2).fill = fill_solid("1E293B")
    ws_sum.cell(r, 2).alignment = Alignment(horizontal="right", vertical="center")
    ws_sum.merge_cells(f"F{r}:G{r}")  # actually use col F
    c2 = ws_sum.cell(r, 6, amt)
    c2.font = Font(name="맑은 고딕", size=12, bold=True, color=color)
    c2.fill = fill_solid("1E293B")
    c2.number_format = '#,##0'
    c2.alignment = Alignment(horizontal="right", vertical="center")
    r += 1


# ═══════════════════════════════════════════════════════
#  시트 4 — 개발 일정 (Schedule)
# ═══════════════════════════════════════════════════════
ws_sched = wb.create_sheet("📅 개발 일정")
ws_sched.sheet_view.showGridLines = False
set_col_width(ws_sched, {"A": 10, "B": 28, "C": 10, "D": 50, "E": 14, "F": 14})
for r in range(1, 40):
    ws_sched.row_dimensions[r].height = 18

# 헤더
ws_sched.row_dimensions[2].height = 34
ws_sched.merge_cells("A2:F2")
cell = ws_sched.cell(2, 1, "개발 일정 — 16주 (4개월) 마일스톤 계획  |  QT-2026-0409-SW-001")
cell.font = Font(name="맑은 고딕", size=13, bold=True, color=C_WHITE)
cell.fill = fill_solid(C_DARK)
cell.alignment = Alignment(horizontal="center", vertical="center")

ws_sched.row_dimensions[3].height = 10
ws_sched.row_dimensions[4].height = 26
sched_headers = ["단계", "구분", "기간", "주요 산출물 및 납품물", "마일스톤", "납부 조건"]
for ci, h in enumerate(sched_headers, 1):
    cell = ws_sched.cell(4, ci, h)
    cell.font = Font(name="맑은 고딕", size=10, bold=True, color=C_WHITE)
    cell.fill = fill_solid(C_NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_thin("334155")

phases = [
    ("Phase 1", "요구사항 분석 · 아키텍처 설계", "1~2주",
     "• BRD/SRS 요구사항 정의서\n• 시스템 아키텍처 설계서\n• Ontology 초안 (OWL 스케치)\n• UI/UX 와이어프레임",
     "M1 — 설계 완료", "계약금 30% (₩24,000,000)",
     "EEF2FF", "4338CA"),
    ("Phase 2", "Ontology + AI 엔진 개발", "3~6주",
     "• OWL/RDF 온톨로지 완성\n• ERP 시맨틱 매핑 엔진\n• 재고 이상 감지 AI 모듈\n• Agentic 경로 결정 알고리즘",
     "M2 — AI 엔진 완성", "중도금 30% (₩24,000,000)",
     "E0F2FE", "0369A1"),
    ("Phase 3", "UI + 드론 제어 개발", "7~10주",
     "• SVG 창고 맵 + 드론 애니메이션\n• ERP 비교 뷰 + Rescan UI\n• 라이브 스캔 피드 로그\n• 일일 보고서 모달",
     "M3 — UI 완성", "중도금 20% (₩16,000,000)",
     "D1FAE5", "065F46"),
    ("Phase 4", "MCP + 보고서 시스템 + DB", "11~13주",
     "• MCP 스케줄러 + 자동 출력\n• 프린터 연동 테스트 완료\n• Admin 보고서 대시보드\n• SQLite 아카이브 DB",
     "M4 — MCP 완성", "—",
     "FEF3C7", "92400E"),
    ("Phase 5", "QA · UAT · 배포 · 교육", "14~16주",
     "• 테스트 완료 보고서\n• 운영자 매뉴얼 최종본\n• 현장 서버 배포 완료\n• 담당자 교육 2회 완료",
     "M5 — 최종 납품", "잔금 20% (₩16,000,000)",
     "FFF1F2", "881337"),
]

r = 5
for ph, name, weeks, deliverables, milestone, payment, bg, fg in phases:
    ws_sched.row_dimensions[r].height = 70
    for c in range(1, 7):
        ws_sched.cell(r, c).fill = fill_solid(bg)
        ws_sched.cell(r, c).border = border_thin()
    ws_sched.cell(r, 1, ph).font = Font(name="맑은 고딕", size=10, bold=True, color=fg)
    ws_sched.cell(r, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws_sched.cell(r, 2, name).font = Font(name="맑은 고딕", size=10, bold=True, color=fg)
    ws_sched.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws_sched.cell(r, 3, weeks).font = Font(name="맑은 고딕", size=10, bold=True, color=fg)
    ws_sched.cell(r, 3).alignment = Alignment(horizontal="center", vertical="center")
    ws_sched.cell(r, 4, deliverables).font = Font(name="맑은 고딕", size=9, color="334155")
    ws_sched.cell(r, 4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    ws_sched.cell(r, 5, milestone).font = Font(name="맑은 고딕", size=9, bold=True, color=fg)
    ws_sched.cell(r, 5).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_sched.cell(r, 6, payment).font = Font(name="맑은 고딕", size=9, bold=True, color="065F46" if "계약금" in payment or "잔금" in payment else "0369A1")
    ws_sched.cell(r, 6).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    r += 1

# 납부 일정 요약
r += 1
ws_sched.row_dimensions[r].height = 24
ws_sched.merge_cells(f"A{r}:F{r}")
cell = ws_sched.cell(r, 1, "납부 일정 요약 (공급가액 ₩80,000,000 기준 · VAT 별도)")
cell.font = Font(name="맑은 고딕", size=11, bold=True, color=C_WHITE)
cell.fill = fill_solid(C_NAVY)
cell.alignment = Alignment(horizontal="center", vertical="center")
r += 1

pay_rows = [
    ("1차 납부", "계약 체결 시 (Phase 1 착수)", "30%", "₩ 24,000,000"),
    ("2차 납부", "M2 완료 시 (AI 엔진 납품)", "30%", "₩ 24,000,000"),
    ("3차 납부", "M3 완료 시 (UI 완성 납품)", "20%", "₩ 16,000,000"),
    ("4차 납부 (잔금)", "M5 완료 시 (최종 납품)", "20%", "₩ 16,000,000"),
]
pay_colors = ["EEF2FF", "E0F2FE", "D1FAE5", "FEF3C7"]
pay_texts  = ["4338CA", "0369A1", "065F46", "92400E"]

for i, (stage, trigger, pct, amt) in enumerate(pay_rows):
    ws_sched.row_dimensions[r].height = 22
    bg = pay_colors[i]
    fg = pay_texts[i]
    vals = [stage, trigger, pct, amt]
    col_map = [1, 2, 3, 6]
    for ci, v in zip(col_map, vals):
        cell = ws_sched.cell(r, ci, v)
        cell.font = Font(name="맑은 고딕", size=10, bold=True, color=fg)
        cell.fill = fill_solid(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin()
    for c in [4, 5]:
        ws_sched.cell(r, c).fill = fill_solid(bg)
        ws_sched.cell(r, c).border = border_thin()
    r += 1


# ═══════════════════════════════════════════════════════
#  시트 5 — 계약 조건 (Terms)
# ═══════════════════════════════════════════════════════
ws_terms = wb.create_sheet("📝 계약 조건")
ws_terms.sheet_view.showGridLines = False
set_col_width(ws_terms, {"A": 5, "B": 28, "C": 55, "D": 5})
for r in range(1, 80):
    ws_terms.row_dimensions[r].height = 18

ws_terms.row_dimensions[2].height = 34
ws_terms.merge_cells("A2:D2")
cell = ws_terms.cell(2, 1, "계약 조건 · 납품 범위 · 보증 조건  |  QT-2026-0409-SW-001")
cell.font = Font(name="맑은 고딕", size=13, bold=True, color=C_WHITE)
cell.fill = fill_solid(C_DARK)
cell.alignment = Alignment(horizontal="center", vertical="center")

terms_sections = [
    ("📦 납품 범위", [
        "소스코드 전체 (GitHub Private Repository 이전)",
        "실행 파일 및 Ubuntu 서버 배포 패키지",
        "REST API 명세서 (Swagger / Markdown)",
        "시스템 아키텍처 문서 및 DB 스키마",
        "운영자 사용 매뉴얼 (한글) 및 설정 가이드",
        "현장 설치 지원 및 운영 교육 2회 (각 2시간)",
    ], "D1FAE5", "065F46"),
    ("🛡️ 보증 및 유지보수", [
        "납품 후 3개월 무상 A/S (버그 수정 포함)",
        "SLA: 장애 발생 시 24시간 이내 대응",
        "원격 지원 월 4회 포함",
        "15레벨 전체 확장 시 추가 견적 협의",
        "삼성 ERP API 정식 연동 시 별도 협의",
        "기능 추가·확장은 별도 협의 처리",
    ], "EEF2FF", "4338CA"),
    ("⚖️ 계약 일반 조건", [
        "견적 유효기간: 발행일로부터 30일 이내",
        "소스코드 저작권: 최종 잔금 납부 완료 후 이전",
        "하드웨어(드론·프린터·서버) 비용 미포함",
        "고객 요구 변경사항은 추가 견적 협의",
        "천재지변·고객 귀책 지연은 납기 조정 협의",
        "비밀유지계약(NDA) 체결 후 개발 착수",
    ], "FEF3C7", "92400E"),
    ("🔧 기술 사양 전제 조건", [
        "창고 내 Wi-Fi 또는 유선 네트워크 제공",
        "Ubuntu 20.04+ 서버 환경 제공",
        "삼성 ERP CSV / API 샘플 데이터 제공",
        "네트워크 프린터 IP · 포트 정보 제공",
        "드론 하드웨어 SDK / API 접근 권한 제공",
        "현장 테스트 환경 및 접근 권한 협조",
    ], "FCE7F3", "9D174D"),
    ("🏗️ 기술 스택 요약", [
        "Backend: Python 3.10+ · Flask · SQLite · threading",
        "Frontend: HTML5 · CSS3 · Vanilla JavaScript · SVG",
        "Ontology: OWL 2.0 · RDF/Turtle · SPARQL 1.1 · RDFLib",
        "AI: Rule-based Engine · NLG · Diff Algorithm · Risk Scoring",
        "MCP: Background Daemon · RAW TCP 9100 · IPP 631 · PJL",
        "DevOps: GitHub · CI/CD · Ubuntu systemd · PM2",
    ], "F5F3FF", "5B21B6"),
    ("✍️ 서명란", [
        "발주처: 삼성반도체 주식회사 DS부문 스마트팩토리팀",
        "담당자 서명: _______________________   (인)",
        "날짜: 2026년   월   일",
        "",
        "공급사: Warehouse AI Solutions",
        "대표이사 서명: _______________________   (인)  / 2026년 4월 9일",
    ], "F8FAFC", "334155"),
]

r = 4
for title, items, bg, fg in terms_sections:
    ws_terms.row_dimensions[r].height = 8
    r += 1
    ws_terms.row_dimensions[r].height = 24
    ws_terms.merge_cells(f"A{r}:D{r}")
    cell = ws_terms.cell(r, 1, f"  {title}")
    cell.font = Font(name="맑은 고딕", size=11, bold=True, color=fg)
    cell.fill = fill_solid(bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(left=Side(style="thick", color=fg), bottom=Side(style="thin", color="E2E8F0"))
    r += 1

    for item in items:
        ws_terms.row_dimensions[r].height = 19
        ws_terms.merge_cells(f"A{r}:D{r}")
        prefix = "  ▸  " if item.strip() else ""
        cell = ws_terms.cell(r, 1, f"{prefix}{item}")
        cell.font = Font(name="맑은 고딕", size=10, color="334155" if item else "FFFFFF")
        cell.fill = fill_solid(bg if item else "FFFFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border(left=Side(style="medium", color=fg), bottom=Side(style="thin", color="F1F5F9"))
        r += 1


# ═══════════════════════════════════════════════════════
#  최종 저장
# ═══════════════════════════════════════════════════════
output_path = "/home/user/webapp/quotation_QT-2026-0409-SW-001.xlsx"
wb.save(output_path)
print(f"✅ Excel 파일 생성 완료: {output_path}")
print(f"   파일 크기: {os.path.getsize(output_path):,} bytes")
